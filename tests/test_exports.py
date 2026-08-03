import io

import pytest
from fastapi.testclient import TestClient

from api.index import app
from tests.conftest import patch_db

client = TestClient(app)
DEV_KEY = 'dev-key-test'
USER_ROW = {'id': 1, 'telegram_id': 123456789, 'username': 'developer', 'first_name': 'Developer',
            'last_name': None, 'is_blocked': False}
MEMBER_ADMIN = {'id': 2, 'workspace_id': 3, 'user_id': 1, 'role': 'admin', 'status': 'active', 'channel_scope': []}
POST_ROW = {'id': 1, 'title': 'Пост', 'text': 'текст', 'status': 'published', 'scheduled_at': None,
            'published_at': None, 'last_error': None, 'channel_title': 'Канал'}
BOOKING_ROW = {'id': 5, 'advertiser_name': 'ООО Реклама', 'channel_title': 'Канал', 'format': 'post',
               'cost': 5000, 'currency': 'RUB', 'status': 'confirmed', 'payment_status': 'paid',
               'publish_at': None, 'erid': 'erid:1', 'erid_required': True}
TX_ROW = {'id': 10, 'type': 'income', 'amount': 1000, 'currency': 'RUB', 'category': 'advertising',
          'description': 'Оплата', 'occurred_at': None}


def auth_headers():
    return {'X-Dev-Api-Key': DEV_KEY}


@pytest.fixture(autouse=True)
def _env(monkeypatch):
    monkeypatch.setenv('DEV_API_KEY', DEV_KEY)
    monkeypatch.setenv('BOT_TOKEN', 'test-token')


def test_export_posts_csv(monkeypatch):
    conn = patch_db(monkeypatch, [USER_ROW, MEMBER_ADMIN, [POST_ROW]])
    monkeypatch.setattr('api.exports.connect', lambda: conn)
    r = client.get('/api/workspaces/3/export/posts?format=csv', headers=auth_headers())
    assert r.status_code == 200
    assert 'text/csv' in r.headers['content-type']
    body = r.content.decode('utf-8-sig')
    assert 'Пост' in body and 'published' in body


def test_export_posts_xlsx(monkeypatch):
    conn = patch_db(monkeypatch, [USER_ROW, MEMBER_ADMIN, [POST_ROW]])
    monkeypatch.setattr('api.exports.connect', lambda: conn)
    r = client.get('/api/workspaces/3/export/posts?format=xlsx', headers=auth_headers())
    assert r.status_code == 200
    assert 'spreadsheetml' in r.headers['content-type']
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.active['A1'].value == 'ID'


def test_export_posts_pdf(monkeypatch):
    conn = patch_db(monkeypatch, [USER_ROW, MEMBER_ADMIN, [POST_ROW]])
    monkeypatch.setattr('api.exports.connect', lambda: conn)
    r = client.get('/api/workspaces/3/export/posts?format=pdf', headers=auth_headers())
    assert r.status_code == 200
    assert 'application/pdf' in r.headers['content-type']
    assert r.content[:4] == b'%PDF'


def test_export_bookings_csv(monkeypatch):
    conn = patch_db(monkeypatch, [USER_ROW, MEMBER_ADMIN, [BOOKING_ROW]])
    monkeypatch.setattr('api.exports.connect', lambda: conn)
    r = client.get('/api/workspaces/3/export/bookings?format=csv', headers=auth_headers())
    assert r.status_code == 200
    assert 'ООО Реклама' in r.content.decode('utf-8-sig')


def test_export_bookings_xlsx(monkeypatch):
    conn = patch_db(monkeypatch, [USER_ROW, MEMBER_ADMIN, [BOOKING_ROW]])
    monkeypatch.setattr('api.exports.connect', lambda: conn)
    r = client.get('/api/workspaces/3/export/bookings?format=xlsx', headers=auth_headers())
    assert r.status_code == 200
    assert 'spreadsheetml' in r.headers['content-type']


def test_export_finance_xlsx(monkeypatch):
    conn = patch_db(monkeypatch, [USER_ROW, MEMBER_ADMIN, [TX_ROW]])
    monkeypatch.setattr('api.exports.connect', lambda: conn)
    r = client.get('/api/workspaces/3/export/finance?format=xlsx', headers=auth_headers())
    assert r.status_code == 200
    assert 'spreadsheetml' in r.headers['content-type']
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.active['B1'].value == 'Тип'


def test_export_finance_pdf(monkeypatch):
    conn = patch_db(monkeypatch, [USER_ROW, MEMBER_ADMIN, [TX_ROW]])
    monkeypatch.setattr('api.exports.connect', lambda: conn)
    r = client.get('/api/workspaces/3/export/finance?format=pdf', headers=auth_headers())
    assert r.status_code == 200
    assert 'application/pdf' in r.headers['content-type']
    assert r.content[:4] == b'%PDF'


def test_export_finance_csv(monkeypatch):
    conn = patch_db(monkeypatch, [USER_ROW, MEMBER_ADMIN, [TX_ROW]])
    monkeypatch.setattr('api.exports.connect', lambda: conn)
    r = client.get('/api/workspaces/3/export/finance?format=csv', headers=auth_headers())
    assert r.status_code == 200
    assert 'Доход' in r.content.decode('utf-8-sig')


def test_create_export_job(monkeypatch):
    job = {'id': 1, 'kind': 'posts', 'format': 'pdf', 'status': 'pending'}
    conn = patch_db(monkeypatch, [USER_ROW, MEMBER_ADMIN, job])
    monkeypatch.setattr('api.exports.connect', lambda: conn)
    r = client.post('/api/workspaces/3/exports', json={'kind': 'posts', 'format': 'pdf'},
                    headers=auth_headers())
    assert r.status_code == 201
    assert r.json()['status'] == 'pending'
    assert 'Telegram' in r.json()['message']
    # в запись попал telegram_id пользователя
    calls = [call for cur in conn.cursors for call in cur.calls]
    insert = next(sql for sql, _ in calls if 'INSERT INTO cd_exports' in sql)
    assert 'telegram_id' in insert


def test_create_finance_export_job_for_period(monkeypatch):
    job = {'id': 2, 'kind': 'finance', 'format': 'xlsx', 'status': 'pending'}
    conn = patch_db(monkeypatch, [USER_ROW, MEMBER_ADMIN, job])
    monkeypatch.setattr('api.exports.connect', lambda: conn)
    r = client.post('/api/workspaces/3/exports',
                    json={'kind': 'finance', 'format': 'xlsx', 'period_year': 2026, 'period_month': 8},
                    headers=auth_headers())
    assert r.status_code == 201
    insert = next(params for cur in conn.cursors for sql, params in cur.calls if 'INSERT INTO cd_exports' in sql)
    assert insert[-2:] == (2026, 8)


def test_create_export_job_bad_kind(monkeypatch):
    conn = patch_db(monkeypatch, [USER_ROW, MEMBER_ADMIN])
    monkeypatch.setattr('api.exports.connect', lambda: conn)
    r = client.post('/api/workspaces/3/exports', json={'kind': 'garbage', 'format': 'csv'},
                    headers=auth_headers())
    assert r.status_code == 422


def test_create_export_job_finance_viewer_denied(monkeypatch):
    viewer = {'id': 2, 'workspace_id': 3, 'user_id': 1, 'role': 'viewer', 'status': 'active', 'channel_scope': []}
    conn = patch_db(monkeypatch, [USER_ROW, viewer])
    monkeypatch.setattr('api.exports.connect', lambda: conn)
    r = client.post('/api/workspaces/3/exports', json={'kind': 'finance', 'format': 'csv'},
                    headers=auth_headers())
    assert r.status_code == 403
