from __future__ import annotations
import csv
import io
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from api.auth import current_user
from api.db import connect
from api.permissions import membership
from api.rbac import require_action

router = APIRouter(prefix='/api', tags=['exports'])

EXPORT_KINDS = {'posts', 'bookings', 'finance', 'media_kits'}
EXPORT_FORMATS = {'csv', 'xlsx', 'pdf'}


class ExportRequest(BaseModel):
    kind: str
    format: str = 'csv'
    period_year: int | None = None
    period_month: int | None = None


def _fmt(dt) -> str:
    if not dt:
        return ''
    if isinstance(dt, datetime):
        return dt.strftime('%Y-%m-%d %H:%M')
    return str(dt)


def _stream_csv(headers: list[str], rows: list[list]) -> StreamingResponse:
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=';')
    writer.writerow(headers)
    writer.writerows(rows)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type='text/csv; charset=utf-8',
                             headers={'Content-Disposition': 'attachment; filename="export.csv"'})


def posts_csv(posts: list[dict]) -> StreamingResponse:
    headers = ['ID', 'Заголовок', 'Текст', 'Статус', 'Канал', 'Запланировано', 'Опубликовано', 'Ошибка']
    rows = [[p['id'], p.get('title') or '', (p.get('text') or '')[:200], p.get('status') or '',
             p.get('channel_title') or '', _fmt(p.get('scheduled_at')), _fmt(p.get('published_at')),
             p.get('last_error') or ''] for p in posts]
    return _stream_csv(headers, rows)


def bookings_csv(bookings: list[dict]) -> StreamingResponse:
    headers = ['ID', 'Рекламодатель', 'Канал', 'Формат', 'Стоимость', 'Валюта', 'Статус', 'Оплата',
               'Публикация', 'ERID', 'ERID требуется']
    rows = [[b['id'], b.get('advertiser_name') or '', b.get('channel_title') or '', b.get('format') or '',
             b.get('cost'), b.get('currency') or '', b.get('status') or '', b.get('payment_status') or '',
             _fmt(b.get('publish_at')), b.get('erid') or '', 'да' if b.get('erid_required', True) else 'нет']
            for b in bookings]
    return _stream_csv(headers, rows)


def finance_csv(transactions: list[dict]) -> StreamingResponse:
    headers = ['ID', 'Тип', 'Сумма', 'Валюта', 'Категория', 'Описание', 'Дата']
    rows = [[t['id'], 'Доход' if t.get('type') == 'income' else 'Расход', t.get('amount'), t.get('currency') or '',
             t.get('category') or '', t.get('description') or '', _fmt(t.get('occurred_at'))]
            for t in transactions]
    return _stream_csv(headers, rows)


def posts_xlsx(posts: list[dict]) -> StreamingResponse:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Публикации'
    ws.append(['ID', 'Заголовок', 'Текст', 'Статус', 'Канал', 'Запланировано', 'Опубликовано', 'Ошибка'])
    for p in posts:
        ws.append([p['id'], p.get('title') or '', (p.get('text') or '')[:200], p.get('status') or '',
                   p.get('channel_title') or '', _fmt(p.get('scheduled_at')), _fmt(p.get('published_at')),
                   p.get('last_error') or ''])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer,
                             media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             headers={'Content-Disposition': 'attachment; filename="posts.xlsx"'})


def finance_xlsx(transactions: list[dict]) -> StreamingResponse:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Финансы'
    ws.append(['ID', 'Тип', 'Сумма', 'Валюта', 'Категория', 'Описание', 'Дата'])
    for t in transactions:
        ws.append([t['id'], 'Доход' if t.get('type') == 'income' else 'Расход', float(t.get('amount') or 0),
                   t.get('currency') or '', t.get('category') or '', t.get('description') or '',
                   _fmt(t.get('occurred_at'))])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer,
                             media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             headers={'Content-Disposition': 'attachment; filename="finance.xlsx"'})


def finance_pdf(transactions: list[dict]) -> StreamingResponse:
    from fpdf import FPDF
    fonts = Path(__file__).resolve().parents[1] / 'assets' / 'fonts'
    pdf = FPDF(orientation='L')
    pdf.add_font('DejaVu', '', str(fonts / 'DejaVuSans.ttf'))
    pdf.add_font('DejaVu', 'B', str(fonts / 'DejaVuSans-Bold.ttf'))
    pdf.add_page()
    pdf.set_font('DejaVu', 'B', 14)
    pdf.cell(0, 10, 'ChannelDesk - Финансы', new_x='LMARGIN', new_y='NEXT')
    headers = ['ID', 'Тип', 'Сумма', 'Валюта', 'Категория', 'Описание', 'Дата']
    widths = [12, 24, 30, 18, 34, 116, 38]
    pdf.set_font('DejaVu', 'B', 8)
    for header, width in zip(headers, widths):
        pdf.cell(width, 7, header, border=1)
    pdf.ln()
    pdf.set_font('DejaVu', '', 8)
    for t in transactions:
        values = [t.get('id'), 'Доход' if t.get('type') == 'income' else 'Расход', t.get('amount'),
                  t.get('currency') or '', t.get('category') or '', t.get('description') or '',
                  _fmt(t.get('occurred_at'))]
        for value, width in zip(values, widths):
            pdf.cell(width, 7, '' if value is None else str(value)[:70], border=1)
        pdf.ln()
    buffer = io.BytesIO()
    pdf.output(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type='application/pdf',
                             headers={'Content-Disposition': 'attachment; filename="finance.pdf"'})


def media_kits_pdf(kits: list[dict]) -> StreamingResponse:
    from fpdf import FPDF
    fonts = Path(__file__).resolve().parents[1] / 'assets' / 'fonts'
    pdf = FPDF()
    pdf.add_font('DejaVu', '', str(fonts / 'DejaVuSans.ttf'))
    pdf.add_font('DejaVu', 'B', str(fonts / 'DejaVuSans-Bold.ttf'))
    if not kits:
        pdf.add_page()
        pdf.set_font('DejaVu', 'B', 16)
        pdf.cell(0, 10, 'ChannelDesk - Медиакиты', new_x='LMARGIN', new_y='NEXT')
        pdf.set_font('DejaVu', '', 10)
        pdf.cell(0, 7, 'Медиакитов пока нет.', new_x='LMARGIN', new_y='NEXT')
    for kit in kits:
        pdf.add_page()
        pdf.set_font('DejaVu', 'B', 18)
        pdf.cell(0, 11, kit.get('name') or 'Медиакит', new_x='LMARGIN', new_y='NEXT')
        pdf.set_font('DejaVu', '', 10)
        pdf.cell(0, 7, f"Канал: {kit.get('channel_title') or 'не указан'}",
                 new_x='LMARGIN', new_y='NEXT')
        pdf.ln(3)
        if kit.get('description'):
            pdf.set_font('DejaVu', 'B', 11)
            pdf.cell(0, 7, 'О канале', new_x='LMARGIN', new_y='NEXT')
            pdf.set_font('DejaVu', '', 10)
            pdf.multi_cell(0, 6, str(kit['description']), new_x='LMARGIN', new_y='NEXT')
        stats = kit.get('stats') or {}
        if stats:
            pdf.set_font('DejaVu', 'B', 11)
            pdf.cell(0, 7, 'Статистика', new_x='LMARGIN', new_y='NEXT')
            pdf.set_font('DejaVu', '', 10)
            for key, value in stats.items():
                pdf.cell(0, 6, f'{key}: {value}', new_x='LMARGIN', new_y='NEXT')
        pricing = kit.get('pricing') or []
        if pricing:
            pdf.set_font('DejaVu', 'B', 11)
            pdf.cell(0, 7, 'Стоимость размещений', new_x='LMARGIN', new_y='NEXT')
            pdf.set_font('DejaVu', '', 10)
            for item in pricing:
                if isinstance(item, dict):
                    fmt = item.get('format') or 'размещение'
                    price = item.get('price', '')
                    currency = item.get('currency') or 'RUB'
                    pdf.cell(0, 6, f'{fmt}: {price} {currency}', new_x='LMARGIN', new_y='NEXT')
                else:
                    pdf.cell(0, 6, str(item), new_x='LMARGIN', new_y='NEXT')
        contacts = kit.get('contacts') or {}
        if contacts:
            pdf.set_font('DejaVu', 'B', 11)
            pdf.cell(0, 7, 'Контакты', new_x='LMARGIN', new_y='NEXT')
            pdf.set_font('DejaVu', '', 10)
            for key, value in contacts.items():
                pdf.cell(0, 6, f'{key}: {value}', new_x='LMARGIN', new_y='NEXT')
    buffer = io.BytesIO()
    pdf.output(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type='application/pdf',
                             headers={'Content-Disposition': 'attachment; filename="media-kits.pdf"'})


def posts_pdf(posts: list[dict]) -> StreamingResponse:
    from fpdf import FPDF
    fonts = Path(__file__).resolve().parents[1] / 'assets' / 'fonts'
    pdf = FPDF()
    pdf.add_font('DejaVu', '', str(fonts / 'DejaVuSans.ttf'))
    pdf.add_font('DejaVu', 'B', str(fonts / 'DejaVuSans-Bold.ttf'))
    pdf.add_page()
    pdf.set_font('DejaVu', 'B', 14)
    pdf.cell(0, 10, 'ChannelDesk - Публикации', new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('DejaVu', 'B', 8)
    pdf.cell(10, 7, 'ID', border=1)
    pdf.cell(60, 7, 'Заголовок', border=1)
    pdf.cell(30, 7, 'Статус', border=1)
    pdf.cell(40, 7, 'Канал', border=1)
    pdf.cell(50, 7, 'Запланировано', border=1, new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('DejaVu', '', 8)
    for p in posts:
        pdf.cell(10, 7, str(p['id']), border=1)
        pdf.cell(60, 7, (p.get('title') or '')[:40], border=1)
        pdf.cell(30, 7, (p.get('status') or ''), border=1)
        pdf.cell(40, 7, (p.get('channel_title') or '')[:25], border=1)
        pdf.cell(50, 7, _fmt(p.get('scheduled_at')), border=1, new_x='LMARGIN', new_y='NEXT')
    buffer = io.BytesIO()
    pdf.output(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type='application/pdf',
                             headers={'Content-Disposition': 'attachment; filename="posts.pdf"'})


# --- Маршруты экспорта ---

@router.get('/workspaces/{workspace_id}/exports')
def list_exports(workspace_id: int, user: dict = Depends(current_user)):
    """Статусы последних заданий экспорта (для диагностики)."""
    member = membership(user['id'], workspace_id)
    require_action(member, 'post.view')
    with connect() as conn, conn.cursor() as cur:
        cur.execute("""SELECT id,kind,format,status,error_text,created_at,completed_at
        FROM cd_exports WHERE workspace_id=%s ORDER BY created_at DESC LIMIT 10""", (workspace_id,))
        return cur.fetchall()


@router.post('/workspaces/{workspace_id}/exports', status_code=201)
def create_export_job(workspace_id: int, payload: ExportRequest, user: dict = Depends(current_user)):
    """Создаёт задание на экспорт: файл сгенерирует и отправит бот прямо в Telegram."""
    member = membership(user['id'], workspace_id)
    action = {'posts': 'post.view', 'bookings': 'booking.view', 'finance': 'finance.view',
              'media_kits': 'media_kit.view'}.get(payload.kind)
    if action is None:
        raise HTTPException(422, 'Неизвестный тип экспорта')
    require_action(member, action)
    if payload.format not in EXPORT_FORMATS:
        raise HTTPException(422, 'Неизвестный формат')
    if payload.kind == 'media_kits' and payload.format != 'pdf':
        raise HTTPException(422, 'Медиакиты экспортируются в PDF')
    if payload.period_year is not None or payload.period_month is not None:
        if payload.kind != 'finance' or payload.period_year is None or payload.period_month is None:
            raise HTTPException(422, 'Период можно указать только для финансового экспорта')
        if not 2000 <= payload.period_year <= 2100:
            raise HTTPException(422, 'Год должен быть от 2000 до 2100')
        if not 1 <= payload.period_month <= 12:
            raise HTTPException(422, 'Месяц должен быть от 1 до 12')
    if payload.kind == 'posts' and payload.format == 'pdf':
        pass  # PDF доступен для постов
    with connect() as conn, conn.cursor() as cur:
        cur.execute("""INSERT INTO cd_exports(workspace_id,user_id,telegram_id,kind,format,period_year,period_month)
        VALUES(%s,%s,%s,%s,%s,%s,%s) RETURNING id,kind,format,status""",
                    (workspace_id, user['id'], user['telegram_id'], payload.kind, payload.format,
                     payload.period_year, payload.period_month))
        row = cur.fetchone()
    return {**row, 'message': 'Файл будет отправлен ботом в Telegram в течение ~30 секунд'}

def _load_posts(workspace_id: int) -> list[dict]:
    with connect() as conn, conn.cursor() as cur:
        cur.execute("""SELECT p.*, c.title AS channel_title FROM cd_posts p
        LEFT JOIN cd_channels c ON c.id=p.channel_id WHERE p.workspace_id=%s ORDER BY p.updated_at DESC""",
                    (workspace_id,))
        return cur.fetchall()


def _load_bookings(workspace_id: int) -> list[dict]:
    with connect() as conn, conn.cursor() as cur:
        cur.execute("""SELECT b.*, a.name AS advertiser_name, c.title AS channel_title
        FROM cd_ad_bookings b LEFT JOIN cd_advertisers a ON a.id=b.advertiser_id
        LEFT JOIN cd_channels c ON c.id=b.channel_id WHERE b.workspace_id=%s ORDER BY b.id DESC""",
                    (workspace_id,))
        return cur.fetchall()


def _load_transactions(workspace_id: int) -> list[dict]:
    with connect() as conn, conn.cursor() as cur:
        cur.execute("""SELECT * FROM cd_finance_transactions WHERE workspace_id=%s
        ORDER BY occurred_at DESC, id DESC""", (workspace_id,))
        return cur.fetchall()


def _load_media_kits(workspace_id: int) -> list[dict]:
    with connect() as conn, conn.cursor() as cur:
        cur.execute("""SELECT mk.*, c.title AS channel_title
        FROM cd_media_kits mk LEFT JOIN cd_channels c ON c.id=mk.channel_id
        WHERE mk.workspace_id=%s AND mk.is_active=true ORDER BY mk.name""", (workspace_id,))
        return cur.fetchall()


@router.get('/workspaces/{workspace_id}/export/posts')
def export_posts(workspace_id: int, format: str = 'csv', user: dict = Depends(current_user)):
    member = membership(user['id'], workspace_id)
    require_action(member, 'post.view')
    posts = _load_posts(workspace_id)
    if format == 'xlsx':
        return posts_xlsx(posts)
    if format == 'pdf':
        return posts_pdf(posts)
    return posts_csv(posts)


@router.get('/workspaces/{workspace_id}/export/bookings')
def export_bookings(workspace_id: int, format: str = 'csv', user: dict = Depends(current_user)):
    member = membership(user['id'], workspace_id)
    require_action(member, 'booking.view')
    bookings = _load_bookings(workspace_id)
    return bookings_csv(bookings) if format != 'xlsx' else _bookings_xlsx(bookings)


@router.get('/workspaces/{workspace_id}/export/finance')
def export_finance(workspace_id: int, format: str = 'csv', user: dict = Depends(current_user)):
    member = membership(user['id'], workspace_id)
    require_action(member, 'finance.view')
    transactions = _load_transactions(workspace_id)
    if format == 'xlsx':
        return finance_xlsx(transactions)
    if format == 'pdf':
        return finance_pdf(transactions)
    return finance_csv(transactions)


@router.get('/workspaces/{workspace_id}/export/media-kits')
def export_media_kits(workspace_id: int, format: str = 'pdf', user: dict = Depends(current_user)):
    member = membership(user['id'], workspace_id)
    require_action(member, 'media_kit.view')
    if format != 'pdf':
        raise HTTPException(422, 'Медиакиты экспортируются в PDF')
    return media_kits_pdf(_load_media_kits(workspace_id))


def _bookings_xlsx(bookings: list[dict]) -> StreamingResponse:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Брони'
    ws.append(['ID', 'Рекламодатель', 'Канал', 'Формат', 'Стоимость', 'Валюта', 'Статус', 'Оплата',
               'Публикация', 'ERID', 'ERID требуется'])
    for b in bookings:
        ws.append([b['id'], b.get('advertiser_name') or '', b.get('channel_title') or '', b.get('format') or '',
                   float(b.get('cost') or 0), b.get('currency') or '', b.get('status') or '',
                   b.get('payment_status') or '', _fmt(b.get('publish_at')), b.get('erid') or '',
                   'да' if b.get('erid_required', True) else 'нет'])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer,
                             media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             headers={'Content-Disposition': 'attachment; filename="bookings.xlsx"'})
